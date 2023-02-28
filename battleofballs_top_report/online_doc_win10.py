#!/usr/bin/python3
# _*_ coding: utf-8 _*_

import datetime
import time
import pyautogui
import win32api
import openpyxl
from wechat import *


def open_url(url):
    """打开网站"""
    if url:
        # 搜索浏览器
        # pyautogui.hotkey('win', 'q')  # 打开"Windows 搜索"菜单
        pyautogui.hotkey('win', 's')  # 打开"Windows 搜索"菜单
        time.sleep(0.5)
        print('搜索浏览器应用')
        write('Microsoft Edge')
        time.sleep(0.5)
        browser_image = 'images/microsoft_edge-win10.png'  # 微软浏览器
        browser_location = pyautogui.locateCenterOnScreen(browser_image, confidence=0.85, minSearchTime=2)
        if browser_location:
            print('打开浏览器')
            pyautogui.moveTo(browser_location, duration=0.5)
            pyautogui.click()
            time.sleep(2)  # 等待浏览器启动
            # 最大化当前窗口
            # Windows11快捷命令变了，如果浏览器已经是全屏打开的，默认新打开的浏览器窗口就是全屏的，如果再次点击此快捷键会将窗口居上放置，会影响后面的输入
            pyautogui.keyDown('win')
            pyautogui.press(['up', 'down', 'up'], interval=0.3)  # 兼容窗口居上放置，而不是全屏的情况
            pyautogui.keyUp('win')
            print('浏览器已打开')
            print(f'输入网址 【{url}】')
            write(url)
            pyautogui.press('enter')
            time.sleep(5)
            online_doc_menu_image = 'images/online_doc_menu-win10.png'
            online_doc_menu_location = pyautogui.locateCenterOnScreen(online_doc_menu_image, confidence=0.85, minSearchTime=2)
            if online_doc_menu_location:
                print(f'网站已打开 【{url}】')
            else:
                print('网站打开失败，继续尝试')
                open_url(url)  # 需要重新调用 open_url() 函数
        else:
            print(f'浏览器打开失败，未找到图标 【{browser_image}】')
            exit(1)
    else:
        print('网站打开失败，网址为空')
        exit(1)


def download_online_doc():
    """下载在线文档"""
    online_doc_menu_image = 'images/online_doc_menu-win10.png'
    online_doc_menu_location = pyautogui.locateCenterOnScreen(online_doc_menu_image, confidence=0.85, minSearchTime=2)
    if online_doc_menu_location:
        print('打开在线文档菜单')
        pyautogui.moveTo(online_doc_menu_location, duration=0.5)
        pyautogui.click()
        time.sleep(1)
        online_doc_download_image = 'images/online_doc_download-win10.png'
        online_doc_download_location = pyautogui.locateCenterOnScreen(online_doc_download_image, confidence=0.85, minSearchTime=2)
        if online_doc_download_location:
            print('下载在线文档')
            pyautogui.moveTo(online_doc_download_location, duration=0.5)
            pyautogui.click()
            time.sleep(2)
            print('在线文档下载完成')
        else:
            print(f'在线文档下载失败，未找到下载图标 【{online_doc_download_image}】')
            exit(1)
    else:
        print(f'在线文档菜单打开失败，未找到菜单图标 【{online_doc_menu_image}】')
        exit(1)


def generate_excel(excel, ocr_top_data):
    """生成本地文档表格"""
    wb = openpyxl.load_workbook(excel)
    ws = wb.active
    print(f'表格第一列的数据：{[i[0].value for i in ws.rows]}')
    ws_row_max = ws.max_row  # 获取表格的最大行
    print(f'表格的最大行：{ws_row_max}')
    data_start_row = 4  # 数据起始行
    if ws_row_max < data_start_row:
        print(f'{excel} 表格最大行数小于4，数据格式不完整，请手动更新下数据格式')

    # 当前段位排行榜第一名段位
    t_top_data = [[key, value] for key, value in ocr_top_data.items()]
    first_name, first_stars = t_top_data[0][0], t_top_data[0][1]  # 当前段位排行榜第一名
    print(f'当前段位排行榜第一名：{first_name}，段位：{first_stars}')

    # 更新段位数据
    current_day = datetime.datetime.now().day  # 当前天数
    current_hour = datetime.datetime.now().hour  # 当前小时数
    data_row = 2 + 2 * current_day  # 更新数据的行，表头有两行
    print(f'更新数据的行：{data_row}')

    data_cols = [col for col in range(2, 16, 2)]  # 更新数据的列范围
    print(f'更新数据的列范围：{data_cols}')
    data_hours = [[0, 1, 2]]  # 更新数据的小时数范围
    for i in range(3, 23, 4):
        data_hours.append([i + j for j in range(4)])
    data_hours.append([23])
    print(f'更新数据的小时数范围：{data_hours}')
    data_hour_index = [index for index, hours in enumerate(data_hours) if current_hour in hours][0]
    print(f'更新数据的小时数的索引：{data_hour_index}')
    data_col = data_cols[data_hour_index]
    print(f'更新数据的列：{data_col}')

    # 写入段位数据
    ws.cell(data_row, data_col, int(first_stars))
    wb.save(excel)


def open_doc(doc):
    """打开本地文档"""
    if doc:
        # 打开本地文档
        win32api.ShellExecute(0, 'open', doc, '', '', 1)
        time.sleep(2)
        wps_home_image = 'images/wps_home-win10.png'
        wps_home_location = pyautogui.locateCenterOnScreen(wps_home_image, confidence=0.85, minSearchTime=2)
        if wps_home_location:
            print(f'本地文档已打开 【{doc}】')
            pyautogui.hotkey('ctrl', 'a')  # 全选本地文档内容
            pyautogui.hotkey('ctrl', 'c')  # 复制文档内容
            time.sleep(1)
            print(f'本地文档已复制')
        else:
            print('本地文档未正常打开，继续尝试')
            open_doc(doc)  # 需要重新调用 open_doc() 函数
    else:
        print(f'本地文档 {doc} 为空，请确认！！！')
        exit(1)


def upload_online_doc():
    """更新在线文档"""
    # 打开在线文档
    wps_home_image = 'images/wps_home-win10.png'
    wps_home_coords = pyautogui.locateOnScreen(wps_home_image, confidence=0.85, minSearchTime=2)
    if wps_home_coords:
        wps_home_location = center_upper(wps_home_coords)  # 计算中下部的坐标
        pyautogui.moveTo(wps_home_location, duration=0.5)
        pyautogui.click()
        time.sleep(1)
        wps_online_doc_image = 'images/wps_online_doc-win10.png'
        wps_online_doc_location = pyautogui.locateCenterOnScreen(wps_online_doc_image, confidence=0.85, minSearchTime=2)
        if wps_online_doc_location:
            print('打开在线文档')
            pyautogui.moveTo(wps_online_doc_location, duration=0.5)
            pyautogui.click(clicks=2)  # 双击鼠标左键
            time.sleep(2)
            wps_home_location = pyautogui.locateCenterOnScreen(wps_home_image, confidence=0.85, minSearchTime=2)
            if wps_home_location:
                print('在线文档已打开')
                pyautogui.hotkey('ctrl', 'a')  # 全选在线文档内容
                pyautogui.hotkey('ctrl', 'v')  # 粘贴覆盖文档内容
                time.sleep(1)
                print('在线文档已粘贴覆盖')
                pyautogui.hotkey('ctrl', 's')  # 保存文档内容，并更新在线文档
                time.sleep(2)
                print(f'在线文档已更新')
            else:
                print('在线文档未正常打开，继续尝试')
                upload_online_doc()  # 需要重新调用 upload_online_doc() 函数
        else:
            print('在线文档未找到，继续尝试')
            upload_online_doc()  # 需要重新调用 upload_online_doc() 函数
    else:
        print('WPS首页未正常显示，无法打开在线文档，请确认！！！')
        exit(1)
